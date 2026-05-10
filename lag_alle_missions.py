import psycopg2
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

GPS_JUMP_MS = 50  # m/s

VEHICLES = [
    ('lauv-fridtjof', 'Fridtjof', '1F5C8B'),
    ('lauv-harald',   'Harald',   'C45200'),
    ('lauv-roald',    'Roald',    '0A7A3E'),
    ('lauv-thor',     'Thor',     '6A2FBF'),
    ('lauv-marie',    'Marie',    '007A7A'),
]

conn = psycopg2.connect(host='localhost', port=5432, dbname='rovdb',
                        user='rovadmin', password='rovpassword')
cur = conn.cursor()

cur.execute(f"""
WITH ordered AS (
    SELECT vehicle, name, campaign, year, timestamp, geom,
        LAG(timestamp) OVER (PARTITION BY vehicle, name, year ORDER BY timestamp) AS prev_ts,
        LAG(geom)      OVER (PARTITION BY vehicle, name, year ORDER BY timestamp) AS prev_geom
    FROM auv_tracks WHERE vessel_transit = false
    -- mission_status hentes per mission nedanfor
),
flagged AS (
    SELECT *,
        CASE WHEN prev_ts IS NULL OR EXTRACT(EPOCH FROM (timestamp - prev_ts)) > 3600
             THEN 1 ELSE 0 END AS new_seg
    FROM ordered
),
segmented AS (
    SELECT *,
        SUM(new_seg) OVER (PARTITION BY vehicle, name, year ORDER BY timestamp) AS segment
    FROM flagged
),
seg_times AS (
    SELECT vehicle, name, campaign, year, segment,
        MIN(timestamp) AS start_time,
        MAX(timestamp) AS end_time,
        COUNT(*)        AS n_points,
        EXTRACT(EPOCH FROM (MAX(timestamp) - MIN(timestamp))) / 60.0 AS duration_min
    FROM segmented
    GROUP BY vehicle, name, campaign, year, segment
),
steps AS (
    SELECT vehicle, name, year, segment,
        ST_Distance(geom::geography, prev_geom::geography) AS dist_m,
        EXTRACT(EPOCH FROM (timestamp - prev_ts))           AS dt_sec
    FROM segmented
    WHERE new_seg = 0 AND prev_geom IS NOT NULL
),
seg_dist AS (
    SELECT vehicle, name, year, segment,
        COALESCE(SUM(dist_m) FILTER (WHERE dt_sec > 0 AND dist_m / dt_sec <= {GPS_JUMP_MS}) / 1000.0, 0) AS distance_km,
        COUNT(*) FILTER (WHERE dt_sec > 0 AND dist_m / dt_sec > {GPS_JUMP_MS}) AS n_jumps
    FROM steps
    GROUP BY vehicle, name, year, segment
)
SELECT
    st.vehicle, st.name, st.campaign, st.year, st.segment,
    st.start_time, st.end_time, st.duration_min, st.n_points,
    COALESCE(sd.distance_km, 0) AS distance_km,
    COALESCE(sd.n_jumps, 0)     AS n_jumps,
    COUNT(*) OVER (PARTITION BY st.vehicle, st.name, st.year) AS total_segments,
    st.duration_min < 3 AS aborted,
    ms.mission_status
FROM seg_times st
LEFT JOIN seg_dist sd USING (vehicle, name, year, segment)
LEFT JOIN (
    SELECT DISTINCT vehicle, name, year, mission_status
    FROM auv_tracks WHERE mission_status IS NOT NULL
) ms USING (vehicle, name, year)
ORDER BY st.vehicle, st.year, st.start_time
""")
all_rows = cur.fetchall()
cur.close()
conn.close()

# Grupper per farkost
from collections import defaultdict
by_vehicle = defaultdict(list)
for row in all_rows:
    by_vehicle[row[0]].append(row)

wb = openpyxl.Workbook()
wb.remove(wb.active)  # fjern default-arket

thin = Side(style='thin', color='CCCCCC')
border = Border(left=thin, right=thin, top=thin, bottom=thin)
center = Alignment(horizontal='center', vertical='center')
left   = Alignment(horizontal='left',   vertical='center')

HEADERS = [
    'Namn (mission)', 'Kampanje', 'År',
    'Startdato', 'Starttid', 'Sluttdato', 'Slutttid',
    'Varighet (min)', 'Varighet (t:mm)', 'Distanse (km)',
    'Antall punkter', 'Avbrot-status', 'Mission-status', 'Merknad',
]
COL_WIDTHS = [45, 40, 6, 12, 10, 12, 10, 16, 14, 14, 16, 12, 14, 40]

# Farge per mission_status
STATUS_FILLS = {
    'success':          PatternFill(start_color='D6F0D6', end_color='D6F0D6', fill_type='solid'),  # grøn
    'technical':        PatternFill(start_color='FFE4B5', end_color='FFE4B5', fill_type='solid'),  # oransje
    'user_depth_late':  PatternFill(start_color='D6E8FA', end_color='D6E8FA', fill_type='solid'),  # blå
    'user_depth_early': PatternFill(start_color='EBF3FF', end_color='EBF3FF', fill_type='solid'),  # lysblå
    'user_surface':     PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid'),  # grå
}

for vehicle_id, label, hdr_hex in VEHICLES:
    rows = by_vehicle.get(vehicle_id, [])
    ws = wb.create_sheet(title=label)

    header_font  = Font(bold=True, color='FFFFFF')
    header_fill  = PatternFill(start_color=hdr_hex, end_color=hdr_hex, fill_type='solid')
    aborted_fill = PatternFill(start_color='F5E0E0', end_color='F5E0E0', fill_type='solid')
    split_fill   = PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid')

    ws.append(HEADERS)
    for col_idx, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font, cell.fill, cell.alignment, cell.border = header_font, header_fill, center, border

    ok_count = 0
    for data_row in rows:
        _, name, campaign, year, segment, start_time, end_time, \
            duration_min, n_points, distance_km, n_jumps, total_segments, aborted, mission_status = data_row

        dur = float(duration_min)
        dur_str = f'{int(dur//60)}:{int(dur%60):02d}'

        abort_status = 'Aborted' if aborted else 'OK'
        notes = []
        if total_segments > 1:
            notes.append(f'Del {segment} av {total_segments} (splittet pga. 24t gap)')
        if n_jumps > 0:
            notes.append(f'{n_jumps} GPS-hopp filtrert')

        values = [
            name, campaign, year,
            start_time.strftime('%Y-%m-%d') if start_time else '',
            start_time.strftime('%H:%M:%S') if start_time else '',
            end_time.strftime('%Y-%m-%d')   if end_time   else '',
            end_time.strftime('%H:%M:%S')   if end_time   else '',
            round(dur, 1), dur_str,
            round(float(distance_km), 2),
            n_points, abort_status, mission_status or '', '; '.join(notes),
        ]

        excel_row = ws.max_row + 1
        if aborted:
            fill = aborted_fill
        elif total_segments > 1:
            fill = split_fill
        elif mission_status in STATUS_FILLS:
            fill = STATUS_FILLS[mission_status]
        else:
            fill = None

        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=excel_row, column=col_idx, value=val)
            if fill:
                cell.fill = fill
            cell.border = border
            cell.alignment = left if col_idx in (1, 2, 14) else center

        if not aborted:
            ok_count += 1

    # Totalrad (kun ikke-aborted missions)
    ok_rows = [r for r in rows if not r[12]]  # r[12] = aborted
    last_data = ws.max_row
    total_row = last_data + 2
    totals = [
        (1, f'TOTAL (ikke-aborted missions: {len(ok_rows)})'),
        (8,  round(sum(float(r[7]) for r in ok_rows), 1)),
        (10, round(sum(float(r[9]) for r in ok_rows), 2)),
        (11, sum(r[8] for r in ok_rows)),
    ]
    for col, val in totals:
        c = ws.cell(row=total_row, column=col, value=val)
        c.font = Font(bold=True)

    # Statistikk for mission_status (berre for Fridtjof der vi har data)
    ab_rows = [r for r in rows if r[12]]
    ab_info = ws.cell(row=total_row + 1, column=1,
        value=f'Aborted missions (< 3 min): {len(ab_rows)} stk — ikke med i totalen')
    ab_info.font = Font(italic=True, color='888888')
    if any(r[13] for r in rows):
        from collections import Counter
        status_counts = Counter(r[13] or 'ukjend' for r in rows if not r[12])
        info = '  |  '.join(f'{k}: {v}' for k,v in sorted(status_counts.items()))
        ms_cell = ws.cell(row=total_row + 2, column=1,
            value=f'Mission-status (ikkje-aborted): {info}')
        ms_cell.font = Font(italic=True, color='555555')

    for i, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:{get_column_letter(len(HEADERS))}1'

    print(f'{label}: {len(ok_rows)} OK + {len(ab_rows)} aborted = {len(rows)} totalt')

outpath = '/Users/martinludvigsen/rov-gis/alle_missions.xlsx'
wb.save(outpath)
print(f'\nLagret: {outpath}')
