import psycopg2
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

GPS_JUMP_THRESHOLD_MS = 50  # m/s – over dette regnes som GPS-hopp

conn = psycopg2.connect(host="localhost", port=5432, dbname="rovdb", user="rovadmin", password="rovpassword")
cur = conn.cursor()

cur.execute(f"""
WITH ordered AS (
    SELECT name, campaign, year, timestamp, geom,
        LAG(timestamp) OVER (PARTITION BY name, year ORDER BY timestamp) as prev_ts,
        LAG(geom)      OVER (PARTITION BY name, year ORDER BY timestamp) as prev_geom
    FROM auv_tracks
    WHERE vehicle = 'lauv-fridtjof' AND vessel_transit = false
),
flagged AS (
    SELECT *,
        CASE WHEN prev_ts IS NULL
                  OR EXTRACT(EPOCH FROM (timestamp - prev_ts)) > 3600
             THEN 1 ELSE 0 END as new_seg
    FROM ordered
),
segmented AS (
    SELECT *,
        SUM(new_seg) OVER (PARTITION BY name, year ORDER BY timestamp) as segment
    FROM flagged
),
-- Steg mellom påfølgende punkter innenfor samme segment
steps AS (
    SELECT name, campaign, year, segment,
        ST_Distance(geom::geography, prev_geom::geography)           as dist_m,
        EXTRACT(EPOCH FROM (timestamp - prev_ts))                     as dt_sec
    FROM segmented
    WHERE new_seg = 0 AND prev_geom IS NOT NULL
),
-- Summer kun steg der hastigheten er <= terskel (filtrer GPS-hopp)
valid_steps AS (
    SELECT name, campaign, year, segment, dist_m
    FROM steps
    WHERE dt_sec > 0 AND dist_m / dt_sec <= {GPS_JUMP_THRESHOLD_MS}
),
seg_dist AS (
    SELECT name, campaign, year, segment,
        COALESCE(SUM(dist_m) / 1000.0, 0) as distance_km
    FROM valid_steps
    GROUP BY name, campaign, year, segment
),
seg_times AS (
    SELECT name, campaign, year, segment,
        MIN(timestamp) as start_time,
        MAX(timestamp) as end_time,
        COUNT(*)        as n_points,
        EXTRACT(EPOCH FROM (MAX(timestamp) - MIN(timestamp))) / 60.0 as duration_min
    FROM segmented
    GROUP BY name, campaign, year, segment
),
-- Tell opp antall GPS-hopp per segment
jump_counts AS (
    SELECT name, year, segment,
        COUNT(CASE WHEN dt_sec > 0 AND dist_m / dt_sec > {GPS_JUMP_THRESHOLD_MS} THEN 1 END) as n_jumps
    FROM steps
    GROUP BY name, year, segment
)
SELECT
    st.name, st.campaign, st.year, st.segment,
    st.start_time, st.end_time, st.duration_min, st.n_points,
    COALESCE(sd.distance_km, 0) as distance_km,
    COALESCE(jc.n_jumps, 0)     as n_jumps,
    COUNT(*) OVER (PARTITION BY st.name, st.year) as total_segments
FROM seg_times st
LEFT JOIN seg_dist    sd USING (name, campaign, year, segment)
LEFT JOIN jump_counts jc USING (name, year, segment)
ORDER BY st.year, st.start_time
""")

rows = cur.fetchall()
cur.close()
conn.close()

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "lauv-fridtjof missions"

header_font  = Font(bold=True, color="FFFFFF")
header_fill  = PatternFill(start_color="1F5C8B", end_color="1F5C8B", fill_type="solid")
alt_fill     = PatternFill(start_color="E8F0F7", end_color="E8F0F7", fill_type="solid")
split_fill   = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
center_align = Alignment(horizontal="center", vertical="center")
left_align   = Alignment(horizontal="left",   vertical="center")
thin   = Side(style="thin", color="CCCCCC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

headers = [
    "Navn (mission)", "Kampanje", "År",
    "Startdato", "Starttid", "Sluttdato", "Slutttid",
    "Varighet (min)", "Varighet (t:mm)", "Distanse (km)",
    "Antall punkter", "Merknad",
]
ws.append(headers)
for col_idx, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx)
    cell.font, cell.fill, cell.alignment, cell.border = header_font, header_fill, center_align, border

for row_idx, row in enumerate(rows, 2):
    name, campaign, year, segment, start_time, end_time, duration_min, n_points, distance_km, n_jumps, total_segments = row
    dur = float(duration_min)
    dur_str = f"{int(dur//60)}:{int(dur%60):02d}"

    notes = []
    if total_segments > 1:
        notes.append(f"Del {segment} av {total_segments} (splittet pga. 24t gap)")
    if n_jumps > 0:
        notes.append(f"{n_jumps} GPS-hopp filtrert")
    merknad = "; ".join(notes)

    values = [
        name, campaign, year,
        start_time.strftime("%Y-%m-%d") if start_time else "",
        start_time.strftime("%H:%M:%S") if start_time else "",
        end_time.strftime("%Y-%m-%d")   if end_time   else "",
        end_time.strftime("%H:%M:%S")   if end_time   else "",
        round(dur, 1), dur_str,
        round(float(distance_km), 2),
        n_points, merknad,
    ]

    fill = split_fill if total_segments > 1 else (alt_fill if row_idx % 2 == 0 else None)
    for col_idx, val in enumerate(values, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        if fill:
            cell.fill = fill
        cell.border = border
        cell.alignment = left_align if col_idx in (1, 2, 12) else center_align

col_widths = [45, 40, 6, 12, 10, 12, 10, 16, 14, 14, 16, 40]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

last_row = len(rows) + 2
totals = [
    (1, "TOTAL"),
    (8,  round(sum(float(r[6]) for r in rows), 1)),
    (10, round(sum(float(r[8]) for r in rows), 2)),
    (11, sum(r[7] for r in rows)),
]
for col, val in totals:
    ws.cell(row=last_row + 1, column=col, value=val).font = Font(bold=True)

outpath = "/Users/martinludvigsen/rov-gis/fridtjof_missions.xlsx"
wb.save(outpath)

total_jumps = sum(r[9] for r in rows)
print(f"Lagret: {outpath}")
print(f"Antall rader: {len(rows)}")
print(f"GPS-hopp filtrert totalt: {total_jumps} steg på tvers av {sum(1 for r in rows if r[9]>0)} missions")
print(f"Total distanse (renset): {round(sum(float(r[8]) for r in rows), 2)} km")
print(f"Total varighet: {round(sum(float(r[6]) for r in rows)/60, 1)} timer")
